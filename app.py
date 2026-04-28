"""
BWIC Manager — marimo reactive application for two-round CLO BWIC auctions.

Stage 1 (this file): BWIC setup, line builder, Intex upload, multi-scenario pricing table.

Run with::

    marimo edit app.py        # interactive editor
    marimo run  app.py        # read-only app server
"""
from __future__ import annotations

import marimo as mo

app = mo.App(width="full", app_title="BWIC Manager")


# ============================================================================
# Imports
# ============================================================================

@app.cell
def _imports():
    import sys
    import io
    sys.path.insert(0, "src")

    import marimo as mo
    import pandas as pd
    from datetime import date, datetime, timedelta

    from graam.bwic_workflow import Bwic, BwicLine as WfBwicLine, LineState, BidError
    from graam.bwic_pricing import price_bwic_line_multi, BwicLineMulti
    from graam.analytics import BondAnalytics
    from graam.intex_loader import load_intex_excel
    from graam.models import CashflowEntry, CashflowStream
    from graam.enums import Compounding, Frequency

    return (
        BidError,
        BondAnalytics,
        Bwic,
        BwicLineMulti,
        CashflowEntry,
        CashflowStream,
        Compounding,
        Frequency,
        LineState,
        WfBwicLine,
        date,
        datetime,
        io,
        load_intex_excel,
        mo,
        pd,
        price_bwic_line_multi,
        timedelta,
    )


# ============================================================================
# Application state
# ============================================================================

@app.cell
def _state(mo):
    get_state, set_state = mo.state(
        {
            "bwic_id": "BWIC-2026-04-28-001",
            "seller": "ABC Asset Mgmt",
            "lines_config": [],   # list of dicts: line_id, cusip, tranche, face, nc_date, reserve, market_stream, base_stream
            "bwic": None,         # graam.bwic_workflow.Bwic — created on Open R1
            "phase": "setup",     # setup | r1_open | r1_closed | r2_open | awarded
            "notifications": [],
            "r1_end_time": None,  # datetime — when R1 should close (for countdown display)
            "r2_end_time": None,
        }
    )
    return get_state, set_state


# ============================================================================
# Helper functions
# ============================================================================

@app.cell
def _helpers(BwicLineMulti, io, load_intex_excel, price_bwic_line_multi):

    def load_stream_from_bytes(data: bytes, settle_date=None, balance_override=None):
        """Load a CashflowStream from uploaded Excel bytes."""
        return load_intex_excel(
            io.BytesIO(data),
            settle_date=settle_date,
            balance_override=balance_override,
        )

    def price_line_at(line_cfg: dict, bid_price: float):
        """Price a line config at the given bid under all available scenarios."""
        streams = {}
        if line_cfg.get("market_stream") is not None:
            streams["Market"] = line_cfg["market_stream"]
        if line_cfg.get("base_stream") is not None:
            streams["Base"] = line_cfg["base_stream"]
        if not streams:
            return None
        return price_bwic_line_multi(
            BwicLineMulti(
                cusip=line_cfg.get("cusip", ""),
                tranche_name=line_cfg.get("tranche", ""),
                bid_price=bid_price,
                nc_date=line_cfg.get("nc_date"),
                streams=streams,
            )
        )

    def fmt(v, decimals=1, suffix=""):
        return f"{v:.{decimals}f}{suffix}" if v is not None else "—"

    def resolve_dealer(dropdown_val: str, other_val: str) -> str:
        """Pick 'Other...' → custom text, else the dropdown value."""
        if dropdown_val == "Other...":
            return (other_val or "").strip()
        return (dropdown_val or "").strip()

    def bid_sanity(bwic, line_id, price, market_spread_bps, ba):
        """Return list of warning strings for a candidate bid (empty if clean)."""
        warnings = []
        if bwic is None:
            return warnings
        try:
            line = bwic.get_line(line_id)
        except Exception:
            return warnings
        # Compare to running median of best bids on this line
        best = line._best_bid_per_dealer()
        prices = sorted(b.price for b in best.values())
        if prices:
            mid = prices[len(prices) // 2]
            if abs(price - mid) > 0.50:
                warnings.append(
                    f"Bid is {abs(price - mid):.2f} pts from running median ({mid:.4f}) — typo?"
                )
        # Compare DM to market spread
        if market_spread_bps and ba is not None:
            try:
                dm = ba.dm_from_price(price)
                if abs(dm - market_spread_bps) > 200:
                    warnings.append(
                        f"DM {dm:.0f} is {abs(dm-market_spread_bps):.0f} bps off market ({market_spread_bps:.0f})"
                    )
            except Exception:
                pass
        return warnings

    def countdown(end_time, now):
        """Return (text, kind) for a countdown badge."""
        if end_time is None:
            return ("Timer not set", "neutral")
        secs = int((end_time - now).total_seconds())
        if secs <= 0:
            over = -secs
            mm, ss = divmod(over, 60)
            return (f"⏰ EXPIRED  ({mm:02d}:{ss:02d} over)", "danger")
        mm, ss = divmod(secs, 60)
        kind = "warn" if secs <= 60 else ("warn" if secs <= 300 else "success")
        return (f"⏳ {mm:02d}:{ss:02d}  remaining (closes {end_time.strftime('%H:%M:%S')})", kind)

    return bid_sanity, countdown, fmt, load_stream_from_bytes, price_line_at, resolve_dealer


# ============================================================================
# Demo cashflow builder (synthetic AAA — usable without uploading any file)
# ============================================================================

@app.cell
def _demo_builder(CashflowEntry, CashflowStream, Compounding, Frequency, date):

    def build_demo_stream(spread_bps=115, reinvest_months=48, amort_months=24):
        settle = date(2026, 4, 28)
        balance = 100_000_000.0
        n = reinvest_months + amort_months
        sofr = [4.00] * 6 + [3.75] * 12 + [3.50] * (n - 18)
        cashflows = []
        bal = balance
        for i in range(1, n + 1):
            yr = settle.year + (settle.month + i - 1) // 12
            mo_ = (settle.month + i - 1) % 12 + 1
            cf_date = date(yr, mo_, 25)
            sofr_pct = sofr[i - 1]
            interest = bal * (sofr_pct + spread_bps / 100.0) / 100.0 / 12.0
            principal = 0.0 if i <= reinvest_months else balance / amort_months
            prev_bal = bal
            bal = max(bal - principal, 0.0)
            cashflows.append(
                CashflowEntry(
                    cashflow_date=cf_date,
                    interest=interest,
                    principal=principal,
                    balance=bal,
                    prev_balance=prev_bal,
                    index_value=sofr_pct,
                )
            )
        return CashflowStream(
            cashflows=cashflows,
            settle_date=settle,
            balance=balance,
            day_counter_name="Actual360",
            compounding=Compounding.SIMPLE,
            frequency=Frequency.MONTHLY,
        )

    return (build_demo_stream,)


# ============================================================================
# Header / status banner
# ============================================================================

@app.cell
def _header(get_state, mo):
    _s = get_state()
    _phase_map = {
        "setup":     ("SETUP",            "neutral"),
        "r1_open":   ("ROUND 1 — OPEN",   "success"),
        "r1_closed": ("ROUND 1 — CLOSED", "warn"),
        "r2_open":   ("ROUND 2 — OPEN",   "success"),
        "r2_closed": ("ROUND 2 — CLOSED", "warn"),
        "awarded":   ("AWARDED",          "success"),
    }
    _label, _kind = _phase_map.get(_s["phase"], ("?", "neutral"))
    _header_md = mo.vstack(
        [
            mo.md("# BWIC Manager"),
            mo.callout(
                mo.md(
                    f"**{_label}** · {len(_s['lines_config'])} line(s) · "
                    f"`{_s['bwic_id']}` · seller _{_s['seller']}_"
                ),
                kind=_kind,
            ),
        ]
    )
    _header_md
    return


# ============================================================================
# Save / Load BWIC state (pickle — preserves cashflow streams + Bwic object)
# ============================================================================

@app.cell
def _save_load(get_state, mo, set_state):
    import pickle, base64

    _s = get_state()
    _bwic_id_safe = (_s.get("bwic_id") or "bwic").replace(" ", "_")

    # Serialize whole state (minus notifications + transient timers)
    _to_save = {k: v for k, v in _s.items() if k not in ("notifications",)}
    try:
        _blob = pickle.dumps(_to_save)
    except Exception:
        _blob = b""

    save_dl = mo.download(
        data=_blob,
        filename=f"{_bwic_id_safe}.bwic.pkl",
        label="↓ Save BWIC state",
    )

    load_upload = mo.ui.file(filetypes=[".pkl"], label="↑ Load BWIC state (.pkl)")

    def _do_load(_):
        if not load_upload.value:
            return
        try:
            loaded = pickle.loads(load_upload.value[0].contents)
            # merge with current notifications cleared
            set_state({**loaded, "notifications": []})
        except Exception as exc:
            set_state({**get_state(),
                       "notifications": get_state()["notifications"] + [f"Load failed: {exc}"]})

    load_btn = mo.ui.button(label="Apply loaded state", kind="success", on_click=_do_load)
    return load_btn, load_upload, save_dl


# ============================================================================
# Round duration inputs + 1Hz tick for countdown displays
# ============================================================================

@app.cell
def _round_duration_inputs(mo):
    r1_dur_inp = mo.ui.number(value=30, start=1, stop=240, step=1, label="R1 duration (min)")
    r2_dur_inp = mo.ui.number(value=15, start=1, stop=240, step=1, label="R2 duration (min)")
    return r1_dur_inp, r2_dur_inp


@app.cell
def _ticker(mo):
    # 1-second refresh — display cells that depend on tick.value re-run each second
    tick = mo.ui.refresh(default_interval="1s", options=["1s", "5s", "off"])
    return (tick,)


@app.cell
def _clear_state(mo):
    # Incremented on successful bid submit when "Reset form" checkbox is on.
    # _bid_stable_inputs depends on this so it recreates the inputs (resetting values).
    get_clear, set_clear = mo.state(0)
    return get_clear, set_clear


@app.cell
def _alert_state(mo):
    # Tracks which timer thresholds have already fired an alert this round
    # to avoid re-alerting on every tick.  Stored as set of (round, threshold_secs).
    get_alerts, set_alerts = mo.state(set())
    return get_alerts, set_alerts


# ============================================================================
# Stable form inputs (don't depend on mutable state — values persist)
# ============================================================================

@app.cell
def _meta_inputs(mo):
    bwic_id_input = mo.ui.text(value="BWIC-2026-04-28-001", label="BWIC ID", full_width=True)
    seller_input  = mo.ui.text(value="ABC Asset Mgmt",      label="Seller",  full_width=True)
    return bwic_id_input, seller_input


@app.cell
def _line_inputs(date, mo):
    deal_name_inp     = mo.ui.text(value="", label="Deal name", placeholder="ABC 2024-1A")
    cusip_inp         = mo.ui.text(value="", label="CUSIP")
    rating_inp        = mo.ui.dropdown(
        options=["AAA", "AA", "A", "BBB", "BB", "B", "Equity"],
        value="AAA",
        label="Rating",
    )
    tranche_inp       = mo.ui.text(value="", label="Tranche label", placeholder="A-1, B, C, ...")
    face_inp          = mo.ui.number(value=100_000_000, step=1_000_000, label="Current Face ($)")
    orig_face_inp     = mo.ui.number(value=100_000_000, step=1_000_000, label="Original Face ($)")
    quoted_spread_inp = mo.ui.number(value=115, step=1, start=0, stop=2000, label="Quoted spread (bps)")
    market_spread_inp = mo.ui.number(value=122, step=1, start=0, stop=2000, label="Market spread (bps)")
    settle_inp        = mo.ui.date(value=date(2026, 4, 28), label="Settle date")
    nc_date_inp       = mo.ui.date(value=date(2027, 4, 25), label="NC end")
    reinvest_end_inp  = mo.ui.date(value=date(2030, 4, 25), label="Reinvest end")
    reserve_inp       = mo.ui.number(value=None, start=0, stop=120, step=0.0625, label="Reserve price (optional)")
    market_upload     = mo.ui.file(filetypes=[".xlsx", ".xls"], label="Market scenario (20 CPR / 2 CDR / 30 SEV)")
    base_upload       = mo.ui.file(filetypes=[".xlsx", ".xls"], label="Base scenario (15 CPR / 5 CDR / 50 SEV)")
    return (
        base_upload,
        cusip_inp,
        deal_name_inp,
        face_inp,
        market_spread_inp,
        market_upload,
        nc_date_inp,
        orig_face_inp,
        quoted_spread_inp,
        rating_inp,
        reinvest_end_inp,
        reserve_inp,
        settle_inp,
        tranche_inp,
    )


# ============================================================================
# Action buttons — Setup phase
# ============================================================================

@app.cell
def _setup_actions(
    base_upload,
    build_demo_stream,
    bwic_id_input,
    cusip_inp,
    date,
    datetime,
    deal_name_inp,
    face_inp,
    get_state,
    load_stream_from_bytes,
    market_spread_inp,
    market_upload,
    mo,
    nc_date_inp,
    orig_face_inp,
    quoted_spread_inp,
    r1_dur_inp,
    rating_inp,
    reinvest_end_inp,
    reserve_inp,
    seller_input,
    set_state,
    settle_inp,
    timedelta,
    tranche_inp,
):
    def _add_line(_):
        s = get_state()
        idx = len(s["lines_config"]) + 1
        line_id = f"L{idx}"

        market_stream = None
        base_stream = None
        notif = list(s["notifications"])

        try:
            if market_upload.value:
                market_stream = load_stream_from_bytes(
                    market_upload.value[0].contents,
                    settle_date=settle_inp.value,
                    balance_override=float(face_inp.value or 0) or None,
                )
        except Exception as exc:
            notif.append(f"{line_id}: Market upload failed — {exc}")
        try:
            if base_upload.value:
                base_stream = load_stream_from_bytes(
                    base_upload.value[0].contents,
                    settle_date=settle_inp.value,
                    balance_override=float(face_inp.value or 0) or None,
                )
        except Exception as exc:
            notif.append(f"{line_id}: Base upload failed — {exc}")

        new_line = {
            "line_id": line_id,
            "deal_name": deal_name_inp.value or f"Demo Deal {idx}",
            "cusip": cusip_inp.value or f"DEMO{idx:04d}",
            "rating": rating_inp.value or "AAA",
            "tranche": tranche_inp.value or rating_inp.value or "A-1",
            "face": float(face_inp.value or 0),
            "orig_face": float(orig_face_inp.value or face_inp.value or 0),
            "quoted_spread": float(quoted_spread_inp.value or 0),
            "market_spread": float(market_spread_inp.value or 0),
            "settle_date": settle_inp.value,
            "nc_date": nc_date_inp.value,
            "reinvest_end": reinvest_end_inp.value,
            "reserve": float(reserve_inp.value) if reserve_inp.value else None,
            "market_stream": market_stream,
            "base_stream": base_stream,
        }
        set_state(
            {
                **s,
                "bwic_id": bwic_id_input.value or s["bwic_id"],
                "seller": seller_input.value or s["seller"],
                "lines_config": s["lines_config"] + [new_line],
                "notifications": notif,
            }
        )

    def _load_demo(_):
        """Load a four-tranche demo BWIC: AAA / AA / A / BBB."""
        s = get_state()
        # (rating, tranche_label, face_mm, quoted_bps, market_bps, nc_yrs)
        _tranches = [
            ("AAA", "A-1",  100, 115, 122, 1.0),
            ("AA",  "B",     50, 175, 185, 1.5),
            ("A",   "C",     25, 250, 270, 2.0),
            ("BBB", "D",     20, 380, 410, 2.5),
        ]
        new_lines = list(s["lines_config"])
        for i, (rating, label, face_mm, qs, ms, nc_yr) in enumerate(_tranches, 1):
            mkt = build_demo_stream(spread_bps=qs)
            base = build_demo_stream(spread_bps=qs)
            new_lines.append({
                "line_id":       f"L{len(new_lines) + 1}",
                "deal_name":     "ABC 2024-1A",
                "cusip":         f"00000{i:03d}A",
                "rating":        rating,
                "tranche":       label,
                "face":          face_mm * 1_000_000.0,
                "orig_face":     face_mm * 1_000_000.0,
                "quoted_spread": float(qs),
                "market_spread": float(ms),
                "settle_date":   date(2026, 4, 28),
                "nc_date":       date(2026 + int(nc_yr), 4, 25),
                "reinvest_end":  date(2030, 4, 25),
                "reserve":       None,
                "market_stream": mkt,
                "base_stream":   base,
            })
        set_state({**s, "lines_config": new_lines})

    def _clear_lines(_):
        s = get_state()
        set_state({**s, "lines_config": [], "notifications": []})

    def _open_r1(_):
        # Lazy import inside the callback so the cell doesn't depend on the workflow module
        from graam.bwic_workflow import Bwic, BwicLine as WfBwicLine
        s = get_state()
        if not s["lines_config"]:
            set_state({**s, "notifications": s["notifications"] + ["Cannot open R1 — add at least one line first."]})
            return
        bwic = Bwic(
            bwic_id=bwic_id_input.value or s["bwic_id"],
            seller=seller_input.value or s["seller"],
            bwic_date=date.today(),
        )
        for cfg in s["lines_config"]:
            streams = {}
            if cfg.get("market_stream") is not None:
                streams["Market"] = cfg["market_stream"]
            if cfg.get("base_stream") is not None:
                streams["Base"] = cfg["base_stream"]
            line = WfBwicLine(
                line_id=cfg["line_id"],
                cusip=cfg.get("cusip", ""),
                tranche_name=cfg.get("tranche", ""),
                current_face=float(cfg.get("face") or 0),
                nc_date=cfg.get("nc_date"),
                reserve_price=cfg.get("reserve"),
                streams=streams,
                pricing_scenario="Market" if "Market" in streams else (next(iter(streams)) if streams else "Market"),
            )
            bwic.add_line(line)
        bwic.open_round1()
        _r1_end = datetime.now() + timedelta(minutes=int(r1_dur_inp.value or 30))
        set_state({**s, "bwic": bwic, "phase": "r1_open", "r1_end_time": _r1_end})

    add_line_btn   = mo.ui.button(label="+ Add Line",                      kind="success", on_click=_add_line)
    demo_btn       = mo.ui.button(label="Load Demo (AAA $100M S+115)",     kind="neutral", on_click=_load_demo)
    clear_btn      = mo.ui.button(label="Clear Lines",                     kind="danger",  on_click=_clear_lines)
    open_r1_btn    = mo.ui.button(label="Open Round 1 →",                  kind="success", on_click=_open_r1)
    return add_line_btn, clear_btn, demo_btn, open_r1_btn


# ============================================================================
# Pricing reference table — multi-scenario DM at par for all loaded lines
# ============================================================================

@app.cell
def _pricing_table(BondAnalytics, get_state, mo, pd, price_line_at):
    _s = get_state()
    _rows = []
    for _cfg in _s["lines_config"]:
        _result = price_line_at(_cfg, 100.0)  # at par
        _row = {
            "Line":      _cfg["line_id"],
            "Deal":      _cfg.get("deal_name", ""),
            "CUSIP":     _cfg.get("cusip", ""),
            "Rating":    _cfg.get("rating", ""),
            "Tranche":   _cfg.get("tranche", ""),
            "Face $mm":  (_cfg.get("face") or 0) / 1_000_000,
            "Factor":    round((_cfg.get("face") or 0) / max(_cfg.get("orig_face") or 1, 1), 4),
            "S+":        _cfg.get("quoted_spread"),
            "Mkt+":      _cfg.get("market_spread"),
            "NC":        _cfg.get("nc_date"),
            "RE end":    _cfg.get("reinvest_end"),
            "Reserve":   _cfg.get("reserve"),
        }
        if _result is not None:
            for _label in ("Market", "Base"):
                _r = _result.results.get(_label)
                if _r is None:
                    continue
                _dm_w = _r.dm_to_worst
                _row[f"{_label} DM-Mat"] = round(_r.dm_to_maturity, 1) if _r.dm_to_maturity is not None else None
                _row[f"{_label} DM-Call"] = round(_r.dm_to_call, 1) if _r.dm_to_call is not None else None
                _row[f"{_label} DM-Worst"] = round(_dm_w, 1) if _dm_w is not None else None
                _row[f"{_label} WAL"] = round(_r.wal_to_maturity, 2) if _r.wal_to_maturity is not None else None
                # Modified duration at par using DM-to-Worst
                _stream = _cfg.get(f"{_label.lower()}_stream")
                if _stream is not None and _dm_w is not None and _dm_w > -9999:
                    try:
                        _mdur = BondAnalytics(_stream).modified_duration(100.0, _dm_w)
                        _row[f"{_label} MDur"] = round(_mdur, 2)
                    except Exception:
                        _row[f"{_label} MDur"] = None
        _rows.append(_row)

    _pricing_df = pd.DataFrame(_rows) if _rows else pd.DataFrame()
    pricing_table_md = (
        mo.vstack(
            [
                mo.md("### Pricing reference (DM at par, both scenarios)"),
                mo.ui.table(_pricing_df, selection=None, page_size=20),
            ]
        )
        if not _pricing_df.empty
        else mo.md("_No lines yet — add one above or click_ **Load Demo**.")
    )
    return (pricing_table_md,)


# ============================================================================
# Dealer roster manager (editable in the UI, persisted to dealers.json)
# ============================================================================

@app.cell
def _roster_manager(mo):
    import json as _rjson, os as _ros
    _roster_path = _ros.path.join(_ros.path.dirname(__file__), "dealers.json")

    def _load_roster():
        try:
            with open(_roster_path) as _f:
                return _rjson.load(_f)
        except Exception:
            return ["GS","JPM","MS","BofA","Citi","Barclays","DB","Wells",
                    "BMO","CS","Nomura","Jefferies","Mizuho","RBC","BNP",
                    "SocGen","HSBC","Natixis","MUFG","Other..."]

    def _save_roster(roster):
        with open(_roster_path, "w") as _f:
            _rjson.dump(roster, _f, indent=2)

    _roster_now = _load_roster()

    add_dealer_inp = mo.ui.text(value="", label="New dealer name", placeholder="e.g. Barclays")
    remove_dealer_dd = mo.ui.dropdown(
        options=_roster_now,
        value=_roster_now[0] if _roster_now else None,
        label="Remove dealer",
    )

    def _add_dealer(_):
        name = (add_dealer_inp.value or "").strip()
        if not name:
            return
        r = _load_roster()
        if name not in r:
            # Insert before "Other..." if present
            if "Other..." in r:
                r.insert(r.index("Other..."), name)
            else:
                r.append(name)
            _save_roster(r)

    def _remove_dealer(_):
        name = remove_dealer_dd.value
        if not name:
            return
        r = _load_roster()
        if name in r:
            r.remove(name)
            _save_roster(r)

    _add_dealer_btn    = mo.ui.button(label="Add →", kind="success", on_click=_add_dealer)
    _remove_dealer_btn = mo.ui.button(label="Remove ×", kind="danger", on_click=_remove_dealer)

    roster_manager_md = mo.accordion({
        "👥 Manage dealer roster": mo.vstack([
            mo.md(f"Current roster: {', '.join(_roster_now)}"),
            mo.hstack([add_dealer_inp, _add_dealer_btn], gap=1, justify="start"),
            mo.hstack([remove_dealer_dd, _remove_dealer_btn], gap=1, justify="start"),
            mo.md("_Changes take effect on next form render (reload roster)._"),
        ])
    })
    return (roster_manager_md,)


# ============================================================================
# Setup section layout
# ============================================================================

@app.cell
def _setup_section(
    add_line_btn,
    base_upload,
    bwic_id_input,
    clear_btn,
    cusip_inp,
    deal_name_inp,
    demo_btn,
    face_inp,
    get_state,
    load_btn,
    load_upload,
    market_spread_inp,
    market_upload,
    mo,
    nc_date_inp,
    open_r1_btn,
    orig_face_inp,
    pricing_table_md,
    quoted_spread_inp,
    r1_dur_inp,
    r2_dur_inp,
    rating_inp,
    reinvest_end_inp,
    reserve_inp,
    roster_manager_md,
    save_dl,
    seller_input,
    settle_inp,
    tranche_inp,
):
    _s = get_state()
    mo.stop(_s["phase"] != "setup", mo.md("_Setup is closed — round in progress._"))

    _setup_md = mo.vstack(
        [
            mo.md("## 1 · BWIC details"),
            mo.hstack([bwic_id_input, seller_input], gap=2, justify="start"),
            mo.md("## 2 · Add lines"),
            mo.md("**Identity**"),
            mo.hstack([deal_name_inp, cusip_inp, rating_inp, tranche_inp], gap=1, justify="start"),
            mo.md("**Sizing & spreads**"),
            mo.hstack([face_inp, orig_face_inp, quoted_spread_inp, market_spread_inp], gap=1, justify="start"),
            mo.md("**Dates & reserve**"),
            mo.hstack([settle_inp, nc_date_inp, reinvest_end_inp, reserve_inp], gap=1, justify="start"),
            mo.md("**Cashflows (Intex Excel)**"),
            mo.hstack([market_upload, base_upload], gap=2, justify="start"),
            mo.hstack([add_line_btn, demo_btn, clear_btn], gap=1, justify="start"),
            mo.md("## 3 · Pricing reference"),
            pricing_table_md,
            mo.md("---"),
            mo.md("## 4 · Round timing"),
            mo.hstack([r1_dur_inp, r2_dur_inp], gap=2, justify="start"),
            mo.md("---"),
            mo.md("### Save / Load BWIC state"),
            mo.hstack([save_dl, load_upload, load_btn], gap=2, justify="start"),
            roster_manager_md,
            mo.md("---"),
            mo.hstack([open_r1_btn], justify="end"),
        ]
    )
    _setup_md
    return


# ============================================================================
# Notifications panel (errors, validation messages)
# ============================================================================

@app.cell
def _notifications(get_state, mo):
    _s = get_state()
    mo.stop(not _s["notifications"])
    _notifs_md = mo.callout(
        mo.vstack([mo.md(f"- {n}") for n in _s["notifications"][-8:]]),
        kind="warn",
    )
    _notifs_md
    return


# ============================================================================
# Stage 2 — Stable bid-form inputs (persist across state changes)
# ============================================================================

@app.cell
def _bid_stable_inputs(mo, get_clear):
    _ = get_clear()  # depend on clear counter — inputs recreate at par when triggered
    # Dealer roster loaded from dealers.json; falls back to built-in defaults
    import json as _json, os as _os
    _roster_path = _os.path.join(_os.path.dirname(__file__), "dealers.json")
    try:
        with open(_roster_path) as _f:
            DEALER_ROSTER = _json.load(_f)
    except Exception:
        DEALER_ROSTER = [
            "GS", "JPM", "MS", "BofA", "Citi", "Barclays", "DB", "Wells",
            "BMO", "CS", "Nomura", "Jefferies", "Mizuho", "RBC", "BNP",
            "SocGen", "HSBC", "Natixis", "MUFG", "Other...",
        ]
    dealer_dd = mo.ui.dropdown(
        options=DEALER_ROSTER, value=DEALER_ROSTER[0] if DEALER_ROSTER else "GS",
        label="Dealer",
    )
    dealer_other_inp = mo.ui.text(value="", label="Other dealer (when 'Other...' picked)")
    price_inp  = mo.ui.number(
        value=100.0, step=0.0625, start=50.0, stop=150.0,
        label="Bid Price (% par)",
    )
    auto_clear_chk = mo.ui.checkbox(value=False, label="Reset form after submit")
    return DEALER_ROSTER, auto_clear_chk, dealer_dd, dealer_other_inp, price_inp


# ============================================================================
# Round 1 — bid entry, live DM preview, blotter, close button
# ============================================================================

@app.cell
def _r1_section(
    BidError, BondAnalytics, auto_clear_chk, bid_sanity, countdown,
    datetime, dealer_dd, dealer_other_inp, fmt, get_alerts, get_clear,
    get_state, mo, pd, price_inp, price_line_at, resolve_dealer,
    set_alerts, set_clear, set_state, tick, timedelta,
):
    _ = tick.value  # subscribe to 1Hz tick so the countdown re-renders
    _s = get_state()
    mo.stop(_s["phase"] != "r1_open")
    _now = datetime.now()
    _cd_text, _cd_kind = countdown(_s.get("r1_end_time"), _now)

    # ── Audio/visual alerts at T-5min, T-1min, T-0 ────────────────────────
    _alert_js = mo.Html("")
    if _s.get("r1_end_time"):
        _secs_left = int((_s["r1_end_time"] - _now).total_seconds())
        _fired = get_alerts()
        _thresholds = {300: ("R1", 300), 60: ("R1", 60), 0: ("R1", 0)}
        for _thresh, _key in _thresholds.items():
            if _secs_left <= _thresh and _key not in _fired:
                _fired = _fired | {_key}
                set_alerts(_fired)
                _freq = 880 if _thresh == 0 else (660 if _thresh == 60 else 440)
                _dur  = 600 if _thresh == 0 else 300
                _alert_js = mo.Html(f"""
<script>
(function(){{
  try {{
    var ctx = new (window.AudioContext||window.webkitAudioContext)();
    var o = ctx.createOscillator(); var g = ctx.createGain();
    o.connect(g); g.connect(ctx.destination);
    o.frequency.value = {_freq}; g.gain.value = 0.25;
    o.start(); setTimeout(function(){{ o.stop(); ctx.close(); }}, {_dur});
  }} catch(e) {{}}
}})();
</script>
<style>
@keyframes flash-warn {{ 0%,100%{{background:#fff}} 50%{{background:#ffe082}} }}
.bwic-alert-flash {{ animation: flash-warn 0.5s ease 3; }}
</style>
<div class="bwic-alert-flash" style="padding:4px 8px;border-radius:4px;font-weight:bold;color:#b71c1c">
  ⏰ R1 {'EXPIRED' if _thresh == 0 else f'T-{_thresh//60}min'} alert
</div>""")
                break

    def _extend_r1(_):
        _st = get_state()
        _end = _st.get("r1_end_time") or datetime.now()
        set_state({**_st, "r1_end_time": _end + timedelta(minutes=5)})

    _extend_r1_btn = mo.ui.button(label="+5 min", kind="neutral", on_click=_extend_r1)

    _bwic   = _s["bwic"]
    _lines  = _s["lines_config"]
    _opts   = {c["line_id"]: f"{c['line_id']} · {c.get('tranche','')} {c.get('cusip','')}"
               for c in _lines}

    r1_line_sel = mo.ui.dropdown(options=_opts, label="Line")

    # ── Live DM preview ────────────────────────────────────────────────────
    _sel   = r1_line_sel.value
    _price = price_inp.value
    _cfg   = next((c for c in _lines if c["line_id"] == _sel), None) if _sel else None
    _dm_txt = "Select a line to see live DM"
    if _cfg:
        try:
            _pr = price_line_at(_cfg, _price)
            if _pr:
                _parts = []
                for _lbl in ("Market", "Base"):
                    _r = _pr.results.get(_lbl)
                    if _r and _r.dm_to_worst is not None:
                        _parts.append(
                            f"**{_lbl}** DM-W: {fmt(_r.dm_to_worst)} | "
                            f"DM-M: {fmt(_r.dm_to_maturity)} | "
                            f"WAL: {fmt(_r.wal_to_maturity, 2)}y"
                        )
                _dm_txt = " · ".join(_parts) if _parts else "No streams loaded for this line"
        except Exception as _exc:
            _dm_txt = f"Pricing error: {_exc}"

    # ── DM sensitivity bracket (±1 / ±5 bps around current price) ─────────
    _dm_sens_md = mo.md("")
    if _cfg and _cfg.get("market_stream") is not None:
        try:
            _ba_sens = BondAnalytics(_cfg["market_stream"])
            _dm_cur = _ba_sens.dm_from_price(_price)
            if _dm_cur > -9999:
                _sens_rows = []
                for _d in (-5, -1, 0, 1, 5):
                    _dm_t = _dm_cur + _d
                    _px   = _ba_sens.price_from_dm(_dm_t) if _d != 0 else _price
                    _sens_rows.append({
                        "Δ DM": f"{_d:+d}" if _d != 0 else "→ bid",
                        "DM (bps)": f"{_dm_t:.1f}",
                        "Price": f"{_px:.4f}",
                        "Δ Price": f"{(_px - _price):+.4f}" if _d != 0 else "—",
                    })
                _dm_sens_md = mo.vstack([
                    mo.md("**DM sensitivity (Market, at bid price)**"),
                    mo.ui.table(pd.DataFrame(_sens_rows), selection=None),
                ])
        except Exception:
            pass

    # ── Bid sanity warnings (live as user types price/line) ───────────────
    _warn_lines = []
    if _sel and _bwic and _cfg and _cfg.get("market_stream") is not None:
        try:
            _ba_w = BondAnalytics(_cfg["market_stream"])
            _warn_lines = bid_sanity(_bwic, _sel, _price, _cfg.get("market_spread"), _ba_w)
        except Exception:
            pass

    # ── Callbacks ──────────────────────────────────────────────────────────
    def _submit(_):
        _st = get_state()
        _b  = _st.get("bwic")
        _dealer = resolve_dealer(dealer_dd.value, dealer_other_inp.value)
        if _b is None or not r1_line_sel.value or not _dealer:
            return
        try:
            _b.submit_bid(r1_line_sel.value, _dealer, price_inp.value)
            set_state({**_st, "bwic": _b, "notifications": []})
            if auto_clear_chk.value:
                set_clear(get_clear() + 1)
        except BidError as _e:
            set_state({**_st, "notifications": _st["notifications"] + [str(_e)]})

    def _close_r1(_):
        _st = get_state()
        _b  = _st.get("bwic")
        if _b is None:
            return
        _b.close_round1(top_n=3)
        set_state({**_st, "bwic": _b, "phase": "r1_closed"})

    _submit_btn   = mo.ui.button(label="Submit Bid ▶",  kind="success", on_click=_submit)
    _close_r1_btn = mo.ui.button(label="Close Round 1", kind="warn",    on_click=_close_r1)

    # ── Blotter (live) — with bid index for deletion ───────────────────────
    _bid_rows = []
    _global_bid_idx = 0
    if _bwic:
        for _line in _bwic.lines:
            for _li, _bid in enumerate(_line.bids):
                if _bid.round != 1:
                    _global_bid_idx += 1
                    continue
                _bid_rows.append({
                    "#":      _global_bid_idx,
                    "Line":   _bid.line_id,
                    "Dealer": _bid.dealer,
                    "Price":  round(_bid.price, 4),
                    "Time":   _bid.timestamp.strftime("%H:%M:%S"),
                })
                _global_bid_idx += 1
    _blotter_df = pd.DataFrame(_bid_rows) if _bid_rows else pd.DataFrame(
        columns=["#", "Line", "Dealer", "Price", "Time"])

    _blotter_tbl = mo.ui.table(_blotter_df, selection="single", page_size=30)

    def _delete_bid(_):
        _st = get_state()
        _b  = _st.get("bwic")
        if _b is None or not _blotter_tbl.value or len(_blotter_tbl.value) == 0:
            return
        _row = _blotter_tbl.value.iloc[0]
        _lid = _row["Line"]
        try:
            _tgt_line = _b.get_line(_lid)
            _tgt_idx  = int(_row["#"])
            # Recompute per-line index from global index
            _line_bids = [
                (gi, b) for gi, b in enumerate(
                    sum([list(l.bids) for l in _b.lines], [])
                )
                if b.line_id == _lid
            ]
            _local_idx = next((i for i, (gi, _) in enumerate(_line_bids) if gi == _tgt_idx), None)
            if _local_idx is not None:
                _b.delete_bid_by_idx(_lid, _local_idx)
                set_state({**_st, "bwic": _b, "notifications": []})
        except Exception as _ex:
            set_state({**_st, "notifications": _st["notifications"] + [f"Delete failed: {_ex}"]})

    _delete_btn = mo.ui.button(label="Delete selected bid", kind="danger", on_click=_delete_bid)

    _warn_panel = (
        mo.callout(
            mo.vstack([mo.md(f"⚠️  {w}") for w in _warn_lines]),
            kind="warn",
        )
        if _warn_lines else mo.md("")
    )

    mo.vstack([
        mo.md("## Round 1 — Bid Intake"),
        _alert_js,
        mo.hstack([
            mo.callout(mo.md(_cd_text), kind=_cd_kind),
            _extend_r1_btn, tick,
        ], gap=1, justify="start"),
        mo.hstack([r1_line_sel, dealer_dd, dealer_other_inp, price_inp,
                   auto_clear_chk, _submit_btn], gap=1, justify="start"),
        mo.callout(mo.md(_dm_txt), kind="neutral"),
        _dm_sens_md,
        _warn_panel,
        mo.md("### Live bids"),
        _blotter_tbl,
        mo.hstack([_delete_btn], justify="start"),
        mo.md("---"),
        mo.hstack([_close_r1_btn], justify="end"),
    ])
    return


# ============================================================================
# Round 1 results — advancing dealers + Open R2
# ============================================================================

@app.cell
def _r1_results(datetime, get_state, mo, pd, r2_dur_inp, set_state, timedelta):
    _s = get_state()
    mo.stop(_s["phase"] != "r1_closed")

    _bwic = _s["bwic"]

    def _open_r2(_):
        _st = get_state()
        _b  = _st.get("bwic")
        if _b is None:
            return
        try:
            _b.open_round2()
            _r2_end = datetime.now() + timedelta(minutes=int(r2_dur_inp.value or 15))
            set_state({**_st, "bwic": _b, "phase": "r2_open", "r2_end_time": _r2_end})
        except Exception as _exc:
            set_state({**_st, "notifications": _st["notifications"] + [str(_exc)]})

    _open_r2_btn = mo.ui.button(label="Open Round 2 →", kind="success", on_click=_open_r2)

    _adv_rows = []
    if _bwic:
        for _line in _bwic.lines:
            _best = _line._best_bid_per_dealer(round_filter=1)
            _ranked = sorted(_best.values(), key=lambda b: (-b.price, b.timestamp))
            for _i, _bid in enumerate(_ranked):
                _adv_rows.append({
                    "Line":     _line.line_id,
                    "Tranche":  _line.tranche_name,
                    "Rank":     _i + 1,
                    "Dealer":   _bid.dealer,
                    "R1 Price": round(_bid.price, 4),
                    "Advance?": "✓" if _bid.dealer in _line.advancing_dealers else "",
                })

    _adv_df = pd.DataFrame(_adv_rows) if _adv_rows else pd.DataFrame()

    mo.vstack([
        mo.md("## Round 1 — Results"),
        mo.ui.table(_adv_df, selection=None, page_size=30),
        mo.md("---"),
        mo.hstack([_open_r2_btn], justify="end"),
    ])
    return


# ============================================================================
# WAL sensitivity — re-prices the demo stream under 4 CPR scenarios
# ============================================================================

@app.cell
def _wal_sensitivity(BondAnalytics, build_demo_stream, get_state, mo, pd, price_line_at):
    _s = get_state()
    mo.stop(_s["phase"] not in ("r1_closed", "r2_open", "awarded"))
    _lines = _s["lines_config"]
    mo.stop(not _lines)

    _cpr_scenarios = [
        ("10 CPR (fast)", 10),
        ("15 CPR (base)", 15),
        ("20 CPR (market)", 20),
        ("25 CPR (stress)", 25),
    ]

    _wal_rows = []
    for _cfg in _lines:
        # If this is a demo line (stream was built with build_demo_stream), rebuild at each CPR.
        # For real Intex streams we can only show the already-projected WAL.
        _qs = _cfg.get("quoted_spread", 115) or 115
        _reinvest_months = 48  # default for demo

        _row = {
            "Line":    _cfg["line_id"],
            "Deal":    _cfg.get("deal_name", ""),
            "Rating":  _cfg.get("rating", ""),
            "Tranche": _cfg.get("tranche", ""),
        }

        for _lbl, _cpr in _cpr_scenarios:
            # Approximate amort months from CPR: faster prepay → shorter amort window
            _amort_months = max(6, int(24 * (20 / max(_cpr, 1))))
            try:
                _s_cpr = build_demo_stream(
                    spread_bps=int(_qs),
                    reinvest_months=_reinvest_months,
                    amort_months=_amort_months,
                )
                _ba = BondAnalytics(_s_cpr)
                _w  = _ba.wal()
                _row[f"WAL @ {_cpr}"] = round(_w, 2)
                # DM at par for this scenario
                _dm = _ba.dm_from_price(100.0)
                _row[f"DM @ {_cpr}"] = round(_dm, 1) if _dm > -9999 else None
            except Exception:
                _row[f"WAL @ {_cpr}"] = None
                _row[f"DM @ {_cpr}"] = None

        # Also show actual WAL from uploaded Market stream if available
        _mkt = _cfg.get("market_stream")
        if _mkt is not None:
            try:
                _ba_m = BondAnalytics(_mkt)
                _row["WAL (Intex mkt)"] = round(_ba_m.wal(), 2)
            except Exception:
                _row["WAL (Intex mkt)"] = None

        _wal_rows.append(_row)

    _wal_df = pd.DataFrame(_wal_rows) if _wal_rows else pd.DataFrame()

    mo.accordion({
        "📈 WAL sensitivity (CPR scenarios — demo engine)": (
            mo.vstack([
                mo.md("WAL and DM-at-par estimated via synthetic amortization at each CPR rate.  "
                      "Real Intex streams show actual projected WAL under 'WAL (Intex mkt)'."),
                mo.ui.table(_wal_df, selection=None, page_size=20),
            ])
            if not _wal_df.empty else mo.md("_(no lines)_")
        )
    })
    return


# ============================================================================
# Round 2 — bid entry, live DM preview, blotter, close button
# ============================================================================

@app.cell
def _r2_section(
    BidError, BondAnalytics, auto_clear_chk, bid_sanity, countdown,
    datetime, dealer_dd, dealer_other_inp, fmt, get_alerts, get_clear,
    get_state, mo, pd, price_inp, price_line_at, resolve_dealer,
    set_alerts, set_clear, set_state, tick, timedelta,
):
    _ = tick.value
    _s = get_state()
    mo.stop(_s["phase"] != "r2_open")
    _now = datetime.now()
    _cd_text, _cd_kind = countdown(_s.get("r2_end_time"), _now)

    # ── Audio/visual alerts at T-5min, T-1min, T-0 ────────────────────────
    _alert_js = mo.Html("")
    if _s.get("r2_end_time"):
        _secs_left = int((_s["r2_end_time"] - _now).total_seconds())
        _fired = get_alerts()
        _thresholds = {300: ("R2", 300), 60: ("R2", 60), 0: ("R2", 0)}
        for _thresh, _key in _thresholds.items():
            if _secs_left <= _thresh and _key not in _fired:
                _fired = _fired | {_key}
                set_alerts(_fired)
                _freq = 880 if _thresh == 0 else (660 if _thresh == 60 else 440)
                _dur  = 600 if _thresh == 0 else 300
                _alert_js = mo.Html(f"""
<script>
(function(){{
  try {{
    var ctx = new (window.AudioContext||window.webkitAudioContext)();
    var o = ctx.createOscillator(); var g = ctx.createGain();
    o.connect(g); g.connect(ctx.destination);
    o.frequency.value = {_freq}; g.gain.value = 0.25;
    o.start(); setTimeout(function(){{ o.stop(); ctx.close(); }}, {_dur});
  }} catch(e) {{}}
}})();
</script>
<style>
@keyframes flash-warn {{ 0%,100%{{background:#fff}} 50%{{background:#ffe082}} }}
.bwic-alert-flash {{ animation: flash-warn 0.5s ease 3; }}
</style>
<div class="bwic-alert-flash" style="padding:4px 8px;border-radius:4px;font-weight:bold;color:#b71c1c">
  ⏰ R2 {'EXPIRED' if _thresh == 0 else f'T-{_thresh//60}min'} alert
</div>""")
                break

    def _extend_r2(_):
        _st = get_state()
        _end = _st.get("r2_end_time") or datetime.now()
        set_state({**_st, "r2_end_time": _end + timedelta(minutes=5)})

    _extend_r2_btn = mo.ui.button(label="+5 min", kind="neutral", on_click=_extend_r2)

    _bwic  = _s["bwic"]
    _lines = _s["lines_config"]

    _adv_map: dict[str, list[str]] = {}
    if _bwic:
        for _line in _bwic.lines:
            _adv_map[_line.line_id] = _line.advancing_dealers

    _opts = {c["line_id"]: f"{c['line_id']} · {c.get('tranche','')} (adv: {', '.join(_adv_map.get(c['line_id'], []))})"
             for c in _lines}

    r2_line_sel = mo.ui.dropdown(options=_opts, label="Line")

    # ── Live DM preview ────────────────────────────────────────────────────
    _sel   = r2_line_sel.value
    _price = price_inp.value
    _cfg   = next((c for c in _lines if c["line_id"] == _sel), None) if _sel else None
    _dm_txt = "Select a line to see live DM"
    if _cfg:
        try:
            _pr = price_line_at(_cfg, _price)
            if _pr:
                _parts = []
                for _lbl in ("Market", "Base"):
                    _r = _pr.results.get(_lbl)
                    if _r and _r.dm_to_worst is not None:
                        _parts.append(
                            f"**{_lbl}** DM-W: {fmt(_r.dm_to_worst)} | "
                            f"DM-M: {fmt(_r.dm_to_maturity)} | "
                            f"WAL: {fmt(_r.wal_to_maturity, 2)}y"
                        )
                _dm_txt = " · ".join(_parts) if _parts else "No streams for this line"
        except Exception as _exc:
            _dm_txt = f"Pricing error: {_exc}"

    # ── DM sensitivity bracket ─────────────────────────────────────────────
    _dm_sens_md = mo.md("")
    if _cfg and _cfg.get("market_stream") is not None:
        try:
            _ba_sens = BondAnalytics(_cfg["market_stream"])
            _dm_cur = _ba_sens.dm_from_price(_price)
            if _dm_cur > -9999:
                _sens_rows = []
                for _d in (-5, -1, 0, 1, 5):
                    _dm_t = _dm_cur + _d
                    _px   = _ba_sens.price_from_dm(_dm_t) if _d != 0 else _price
                    _sens_rows.append({
                        "Δ DM": f"{_d:+d}" if _d != 0 else "→ bid",
                        "DM (bps)": f"{_dm_t:.1f}",
                        "Price": f"{_px:.4f}",
                        "Δ Price": f"{(_px - _price):+.4f}" if _d != 0 else "—",
                    })
                _dm_sens_md = mo.vstack([
                    mo.md("**DM sensitivity (Market, at bid price)**"),
                    mo.ui.table(pd.DataFrame(_sens_rows), selection=None),
                ])
        except Exception:
            pass

    # ── Bid sanity warnings ───────────────────────────────────────────────
    _warn_lines = []
    if _sel and _bwic and _cfg and _cfg.get("market_stream") is not None:
        try:
            _ba_w = BondAnalytics(_cfg["market_stream"])
            _warn_lines = bid_sanity(_bwic, _sel, _price, _cfg.get("market_spread"), _ba_w)
        except Exception:
            pass

    # ── Callbacks ──────────────────────────────────────────────────────────
    def _submit_r2(_):
        _st = get_state()
        _b  = _st.get("bwic")
        _dealer = resolve_dealer(dealer_dd.value, dealer_other_inp.value)
        if _b is None or not r2_line_sel.value or not _dealer:
            return
        try:
            _b.submit_bid(r2_line_sel.value, _dealer, price_inp.value)
            set_state({**_st, "bwic": _b, "notifications": []})
            if auto_clear_chk.value:
                set_clear(get_clear() + 1)
        except BidError as _e:
            set_state({**_st, "notifications": _st["notifications"] + [str(_e)]})

    def _close_r2(_):
        _st = get_state()
        _b  = _st.get("bwic")
        if _b is None:
            return
        _b.close_round2()
        set_state({**_st, "bwic": _b, "phase": "awarded"})

    _submit_btn   = mo.ui.button(label="Submit Bid ▶",  kind="success", on_click=_submit_r2)
    _close_r2_btn = mo.ui.button(label="Close Round 2", kind="warn",    on_click=_close_r2)

    # ── R2 blotter with delete ─────────────────────────────────────────────
    _bid_rows = []
    _global_bid_idx = 0
    if _bwic:
        for _line in _bwic.lines:
            for _bi, _bid in enumerate(_line.bids):
                if _bid.round != 2:
                    _global_bid_idx += 1
                    continue
                _bid_rows.append({
                    "#":      _global_bid_idx,
                    "Line":   _bid.line_id,
                    "Dealer": _bid.dealer,
                    "Price":  round(_bid.price, 4),
                    "Time":   _bid.timestamp.strftime("%H:%M:%S"),
                })
                _global_bid_idx += 1
    _blotter_df = pd.DataFrame(_bid_rows) if _bid_rows else pd.DataFrame(
        columns=["#", "Line", "Dealer", "Price", "Time"])

    _blotter_tbl = mo.ui.table(_blotter_df, selection="single", page_size=30)

    def _delete_r2_bid(_):
        _st = get_state()
        _b  = _st.get("bwic")
        if _b is None or not _blotter_tbl.value or len(_blotter_tbl.value) == 0:
            return
        _row = _blotter_tbl.value.iloc[0]
        _lid = _row["Line"]
        try:
            _tgt_idx = int(_row["#"])
            _line_bids = [
                (gi, b) for gi, b in enumerate(
                    sum([list(l.bids) for l in _b.lines], [])
                )
                if b.line_id == _lid
            ]
            _local_idx = next((i for i, (gi, _) in enumerate(_line_bids) if gi == _tgt_idx), None)
            if _local_idx is not None:
                _b.delete_bid_by_idx(_lid, _local_idx)
                set_state({**_st, "bwic": _b, "notifications": []})
        except Exception as _ex:
            set_state({**_st, "notifications": _st["notifications"] + [f"Delete failed: {_ex}"]})

    _delete_btn = mo.ui.button(label="Delete selected bid", kind="danger", on_click=_delete_r2_bid)

    _adv_lines = [f"**{lid}**: {', '.join(adv)}" for lid, adv in _adv_map.items() if adv]

    _warn_panel = (
        mo.callout(
            mo.vstack([mo.md(f"⚠️  {w}") for w in _warn_lines]),
            kind="warn",
        )
        if _warn_lines else mo.md("")
    )

    mo.vstack([
        mo.md("## Round 2 — Bid Intake"),
        _alert_js,
        mo.hstack([
            mo.callout(mo.md(_cd_text), kind=_cd_kind),
            _extend_r2_btn, tick,
        ], gap=1, justify="start"),
        mo.callout(mo.md("Advancing dealers: " + " · ".join(_adv_lines)), kind="neutral"),
        mo.hstack([r2_line_sel, dealer_dd, dealer_other_inp, price_inp,
                   auto_clear_chk, _submit_btn], gap=1, justify="start"),
        mo.callout(mo.md(_dm_txt), kind="neutral"),
        _dm_sens_md,
        _warn_panel,
        mo.md("### R2 bids submitted"),
        _blotter_tbl,
        mo.hstack([_delete_btn], justify="start"),
        mo.md("---"),
        mo.hstack([_close_r2_btn], justify="end"),
    ])
    return


# ============================================================================
# Stage 3 — Award screen + color sheet + bid log + CSV exports
# ============================================================================

@app.cell
def _award_section(get_state, mo, pd, set_state):
    _s = get_state()
    mo.stop(_s["phase"] != "awarded")

    _bwic = _s["bwic"]
    _cfg_by_id = {c["line_id"]: c for c in _s.get("lines_config", [])}

    def _vs_market(dm_bps, market_bps):
        if dm_bps is None or not market_bps:
            return ""
        d = market_bps - dm_bps
        if abs(d) < 0.05:
            return "  (at market)"
        sign = "+" if d > 0 else "−"
        word = "tighter than" if d > 0 else "wider than"
        return f"  ({sign}{abs(d):.1f} bps {word} mkt {market_bps:.0f})"

    # ── Award cards (one per line) ─────────────────────────────────────────
    _cards = []
    for _line in _bwic.lines:
        _aw = _line.award
        if _aw is None:
            continue
        _cfg = _cfg_by_id.get(_line.line_id, {})
        _hdr = (
            f"### {_line.line_id} · {_cfg.get('deal_name', '')} "
            f"{_cfg.get('rating', '')} {_line.tranche_name}"
        )
        if _aw.is_dnt:
            _cards.append(mo.callout(
                mo.md(f"{_hdr}\n\n**DNT** — {_aw.dnt_reason}"),
                kind="danger",
            ))
            continue
        _mkt = _cfg.get("market_spread")
        _award_md = (
            f"{_hdr}  \n"
            f"**🏆 AWARD** · {_aw.award_dealer} @ **{_aw.award_price:.4f}**"
            + (f"  (DM {_aw.spread_at_award:.1f} bps{_vs_market(_aw.spread_at_award, _mkt)})"
               if _aw.spread_at_award is not None else "")
            + "  \n"
            f"**📊 COVER** · "
            + (f"{_aw.cover_dealer} @ {_aw.cover_price:.4f}" if _aw.cover_dealer else "_no cover_")
            + (f"  (DM {_aw.spread_at_cover:.1f} bps{_vs_market(_aw.spread_at_cover, _mkt)})"
               if _aw.spread_at_cover is not None else "")
        )
        _cards.append(mo.callout(mo.md(_award_md), kind="success"))

    # ── Color sheet ────────────────────────────────────────────────────────
    _color_df = _bwic.color_sheet()

    # ── Bid log (audit trail) ──────────────────────────────────────────────
    _bid_log_df = _bwic.bid_log()
    if not _bid_log_df.empty:
        _bid_log_df = _bid_log_df.copy()
        _bid_log_df["timestamp"] = _bid_log_df["timestamp"].apply(
            lambda t: t.strftime("%Y-%m-%d %H:%M:%S") if hasattr(t, "strftime") else str(t)
        )

    # ── CSV downloads ──────────────────────────────────────────────────────
    _bwic_id_safe = (_s.get("bwic_id") or "bwic").replace(" ", "_")
    _color_csv = _color_df.to_csv(index=False).encode("utf-8") if not _color_df.empty else b""
    _bidlog_csv = _bid_log_df.to_csv(index=False).encode("utf-8") if not _bid_log_df.empty else b""

    _color_dl = mo.download(
        data=_color_csv,
        filename=f"{_bwic_id_safe}_color_sheet.csv",
        label="↓ Color sheet (CSV)",
    )
    _bidlog_dl = mo.download(
        data=_bidlog_csv,
        filename=f"{_bwic_id_safe}_bid_log.csv",
        label="↓ Bid log (CSV)",
    )

    # ── Excel export (openpyxl — conditional formatting) ──────────────────
    def _build_excel():
        import io as _io
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, numbers
        from openpyxl.formatting.rule import ColorScaleRule
        wb = openpyxl.Workbook()
        # Sheet 1: Color sheet
        ws1 = wb.active
        ws1.title = "Color Sheet"
        if not _color_df.empty:
            ws1.append(list(_color_df.columns))
            for _, r in _color_df.iterrows():
                ws1.append([r[c] for c in _color_df.columns])
            # Color-scale on dm_cover column if present
            if "dm_cover" in _color_df.columns:
                _col_idx = list(_color_df.columns).index("dm_cover") + 1
                _col_letter = openpyxl.utils.get_column_letter(_col_idx)
                _last_row = len(_color_df) + 1
                ws1.conditional_formatting.add(
                    f"{_col_letter}2:{_col_letter}{_last_row}",
                    ColorScaleRule(
                        start_type="min", start_color="63BE7B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="F8696B",
                    ),
                )
        # Sheet 2: Bid log
        ws2 = wb.create_sheet("Bid Log")
        if not _bid_log_df.empty:
            ws2.append(list(_bid_log_df.columns))
            for _, r in _bid_log_df.iterrows():
                ws2.append([r[c] for c in _bid_log_df.columns])
        # Sheet 3: Award summary
        ws3 = wb.create_sheet("Awards")
        _aw_header = ["Line","Tranche","Award Dealer","Award Price","DM Award","Cover Dealer","Cover Price","DM Cover","DNT"]
        ws3.append(_aw_header)
        for _line in _bwic.lines:
            if _line.award:
                _a = _line.award
                ws3.append([
                    _line.line_id, _line.tranche_name,
                    _a.award_dealer, _a.award_price, _a.spread_at_award,
                    _a.cover_dealer, _a.cover_price, _a.spread_at_cover,
                    "Yes" if _a.is_dnt else "No",
                ])
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    try:
        _excel_bytes = _build_excel()
    except Exception as _xex:
        _excel_bytes = b""
    _excel_dl = mo.download(
        data=_excel_bytes,
        filename=f"{_bwic_id_safe}_bwic.xlsx",
        label="↓ BWIC workbook (XLSX)",
    )

    # ── PDF one-pager (reportlab) ─────────────────────────────────────────
    def _build_pdf():
        import io as _io
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter,
                                leftMargin=0.75*inch, rightMargin=0.75*inch,
                                topMargin=0.75*inch, bottomMargin=0.75*inch)
        styles = getSampleStyleSheet()
        story = []

        # Title
        story.append(Paragraph(f"BWIC Award Sheet — {_s.get('bwic_id','')}", styles["Title"]))
        story.append(Paragraph(f"Seller: {_s.get('seller','')} | Date: {_bwic.bwic_date}", styles["Normal"]))
        story.append(Spacer(1, 0.2*inch))

        # Award table
        _pdf_data = [["Line","Tranche","Face $MM","Award Dealer","Price","DM (bps)","Cover Dealer","Cover Price","DM Cover","vs Mkt"]]
        for _line in _bwic.lines:
            _a = _line.award
            if _a is None:
                continue
            _cf = _cfg_by_id.get(_line.line_id, {})
            _mkt = _cf.get("market_spread")
            _face_mm = f"{_line.current_face/1e6:.0f}"
            if _a.is_dnt:
                _pdf_data.append([_line.line_id, _line.tranche_name, _face_mm,
                                   "DNT", "—", "—", "—", "—", "—", "—"])
            else:
                _vs = ""
                if _a.spread_at_award is not None and _mkt:
                    _d = _mkt - _a.spread_at_award
                    _vs = f"{_d:+.1f}"
                _pdf_data.append([
                    _line.line_id, _line.tranche_name, _face_mm,
                    _a.award_dealer or "—",
                    f"{_a.award_price:.4f}" if _a.award_price else "—",
                    f"{_a.spread_at_award:.1f}" if _a.spread_at_award else "—",
                    _a.cover_dealer or "—",
                    f"{_a.cover_price:.4f}" if _a.cover_price else "—",
                    f"{_a.spread_at_cover:.1f}" if _a.spread_at_cover else "—",
                    _vs,
                ])
        tbl = Table(_pdf_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2D5986")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EEF4FB")]),
            ("ALIGN", (2,1), (-1,-1), "RIGHT"),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph("Confidential — for compliance archive only.", styles["Italic"]))
        doc.build(story)
        return buf.getvalue()

    try:
        _pdf_bytes = _build_pdf()
    except Exception as _pex:
        _pdf_bytes = b""
    _pdf_dl = mo.download(
        data=_pdf_bytes,
        filename=f"{_bwic_id_safe}_award.pdf",
        label="↓ Award one-pager (PDF)",
    )

    # ── Dealer participation stats (append to dealer_stats.jsonl) ─────────
    import os as _dsos, json as _dsjson
    _stats_path = _dsos.path.join(_dsos.path.dirname(__file__), "dealer_stats.jsonl")
    _bwic_id_str = _s.get("bwic_id") or "?"

    def _log_stats(_):
        _st = get_state()
        _b = _st.get("bwic")
        if _b is None:
            return
        _entry = {
            "bwic_id": _bwic_id_str,
            "date": str(_b.bwic_date),
            "dealers": {},
        }
        for _l in _b.lines:
            _best = _l._best_bid_per_dealer()
            for _dlr, _bid in _best.items():
                rec = _entry["dealers"].setdefault(_dlr, {"bids": 0, "won": 0, "covered": 0, "face_bid": 0, "face_won": 0})
                rec["bids"] += 1
                rec["face_bid"] += _l.current_face
                if _l.award and not _l.award.is_dnt:
                    if _l.award.award_dealer == _dlr:
                        rec["won"] += 1
                        rec["face_won"] += _l.current_face
                    elif _l.award.cover_dealer == _dlr:
                        rec["covered"] += 1
        try:
            with open(_stats_path, "a") as _sf:
                _sf.write(_dsjson.dumps(_entry) + "\n")
            set_state({**_st, "notifications": _st["notifications"] + [f"Dealer stats logged to dealer_stats.jsonl"]})
        except Exception as _e:
            set_state({**_st, "notifications": _st["notifications"] + [f"Stats log failed: {_e}"]})

    _log_stats_btn = mo.ui.button(label="Log dealer stats", kind="neutral", on_click=_log_stats)

    # ── Reset BWIC ─────────────────────────────────────────────────────────
    def _reset(_):
        set_state({
            "bwic_id": "BWIC-2026-04-28-001",
            "seller": "ABC Asset Mgmt",
            "lines_config": [],
            "bwic": None,
            "phase": "setup",
            "notifications": [],
        })

    _reset_btn = mo.ui.button(label="↺ New BWIC", kind="danger", on_click=_reset)

    mo.vstack([
        mo.md("## 🏆 BWIC Awarded"),
        mo.vstack(_cards) if _cards else mo.md("_No awards._"),
        mo.md("### Color sheet"),
        mo.ui.table(_color_df, selection=None, page_size=30) if not _color_df.empty
            else mo.md("_(no color)_"),
        mo.md("### Bid log (audit trail)"),
        mo.ui.table(_bid_log_df, selection=None, page_size=50) if not _bid_log_df.empty
            else mo.md("_(no bids)_"),
        mo.md("### Export"),
        mo.hstack([_color_dl, _bidlog_dl, _excel_dl, _pdf_dl], gap=1, justify="start"),
        mo.hstack([_log_stats_btn], justify="start"),
        mo.md("---"),
        mo.hstack([_reset_btn], justify="end"),
    ])
    return


# ============================================================================
# Dealer participation stats viewer
# ============================================================================

@app.cell
def _dealer_stats_view(mo, pd):
    import os as _stos, json as _stjson
    _stats_path = _stos.path.join(_stos.path.dirname(__file__), "dealer_stats.jsonl")

    _all_records = []
    try:
        with open(_stats_path) as _f:
            for _line in _f:
                _line = _line.strip()
                if _line:
                    _all_records.append(_stjson.loads(_line))
    except FileNotFoundError:
        pass

    _stats_rows = []
    if _all_records:
        _agg: dict = {}
        for _rec in _all_records:
            for _dlr, _d in _rec.get("dealers", {}).items():
                r = _agg.setdefault(_dlr, {"Dealer": _dlr, "BWICs": 0, "Lines Bid": 0, "Won": 0, "Covered": 0, "Face Bid $MM": 0.0, "Face Won $MM": 0.0})
                r["BWICs"] += 1
                r["Lines Bid"] += _d.get("bids", 0)
                r["Won"] += _d.get("won", 0)
                r["Covered"] += _d.get("covered", 0)
                r["Face Bid $MM"] += _d.get("face_bid", 0) / 1e6
                r["Face Won $MM"] += _d.get("face_won", 0) / 1e6
        for _dlr, r in _agg.items():
            r["Win Rate"] = f"{100*r['Won']/max(r['Lines Bid'],1):.1f}%"
            r["Face Bid $MM"] = round(r["Face Bid $MM"], 1)
            r["Face Won $MM"] = round(r["Face Won $MM"], 1)
            _stats_rows.append(r)
        _stats_rows.sort(key=lambda x: -x["Won"])

    _stats_df = pd.DataFrame(_stats_rows) if _stats_rows else pd.DataFrame()

    mo.accordion({
        "📊 Dealer participation stats (lifetime)": (
            mo.vstack([
                mo.md(f"_{len(_all_records)} BWIC(s) on record._"),
                mo.ui.table(_stats_df, selection=None, page_size=30),
            ])
            if not _stats_df.empty
            else mo.md("_No stats yet — click 'Log dealer stats' after awarding a BWIC._")
        )
    })
    return


# ============================================================================
# Bloomberg IB messages — copy/paste these into IB chat during the BWIC
# ============================================================================

@app.cell
def _ib_messages(date, datetime, get_state, mo, pd):
    _s = get_state()
    _bwic = _s.get("bwic")
    _phase = _s["phase"]
    _bwic_id = _s.get("bwic_id") or "BWIC"
    _seller = _s.get("seller") or "Seller"

    def _fmt_face(face):
        if face >= 1_000_000:
            return f"{face/1_000_000:.0f}MM"
        return f"{face:,.0f}"

    def _fmt_dt(dt):
        return dt.strftime("%H:%M EST") if dt else "TBD"

    # ── 1. ANNOUNCEMENT (sent before R1 opens — covers full lineup) ────────
    _lines_cfg = _s["lines_config"]
    _today_str = date.today().strftime("%a %m/%d/%y")
    _r1_when = _fmt_dt(_s.get("r1_end_time"))
    _r2_when = _fmt_dt(_s.get("r2_end_time"))

    _announce_lines = [
        f"*** BWIC ANNOUNCEMENT — {_today_str} ***",
        f"Seller:  {_seller}",
        f"BWIC ID: {_bwic_id}",
        "",
        f"R1 due:  {_r1_when}  (top 3 advance, ties move)",
        f"R2 due:  {_r2_when}  (last & best)",
        "Color:   post-trade, T+0",
        "",
        "LINEUP:",
    ]
    for _i, _c in enumerate(_lines_cfg, 1):
        _nc = _c.get("nc_date").strftime("%m/%y") if _c.get("nc_date") else "n/a"
        _re = _c.get("reinvest_end").strftime("%m/%y") if _c.get("reinvest_end") else "n/a"
        _res = f"{_c['reserve']:.4f}" if _c.get("reserve") else "NONE"
        _of = _c.get("orig_face") or _c.get("face") or 0
        _fct = (_c.get("face") or 0) / max(_of, 1)
        _qs = f"S+{_c['quoted_spread']:.0f}" if _c.get("quoted_spread") else ""
        _ms = f"mkt +{_c['market_spread']:.0f}" if _c.get("market_spread") else ""
        _spread_str = " ".join(p for p in (_qs, _ms) if p)
        _deal = _c.get("deal_name", "")
        _rating = _c.get("rating", "")
        _announce_lines.append(
            f"  {_i}) {_deal} {_rating} {_c.get('tranche','')} {_c.get('cusip','')}"
        )
        _announce_lines.append(
            f"     ${_fmt_face(_c.get('face') or 0)} (f={_fct:.3f}) | "
            f"NC {_nc} | RE {_re} | {_spread_str} | Res: {_res}"
        )
    _announce_lines += [
        "",
        "Format: direct via IB, all-in price (% par).",
        "Bids subject to seller discretion. Cover/color out post-trade.",
    ]
    announce_msg = "\n".join(_announce_lines)

    # ── 2. R1 OPEN ─────────────────────────────────────────────────────────
    r1_open_msg = "\n".join([
        f"*** R1 OPEN — {_bwic_id} ***",
        f"Bids due {_r1_when}.",
        "Direct via IB, all-in price.",
        "Top 3 advance to R2 (ties move).",
    ])

    # ── 3. R1 CLOSED / advancers ───────────────────────────────────────────
    _r1_close_lines = [f"*** R1 CLOSED — {_bwic_id} ***", ""]
    if _bwic and _phase in ("r1_closed", "r2_open", "awarded"):
        _r1_close_lines.append("Advancing to R2:")
        for _line in _bwic.lines:
            _r1_close_lines.append(
                f"  {_line.line_id} {_line.tranche_name}: "
                + (", ".join(_line.advancing_dealers) if _line.advancing_dealers else "—")
            )
        _r1_close_lines += ["", f"R2 opens shortly — last & best by {_r2_when}."]
    else:
        _r1_close_lines.append("(R1 not yet closed)")
    r1_closed_msg = "\n".join(_r1_close_lines)

    # ── 4. R2 OPEN ─────────────────────────────────────────────────────────
    r2_open_msg = "\n".join([
        f"*** R2 OPEN — Last & Best ***",
        f"You are advancing.  Last & best due {_r2_when}.",
        "You may stay flat — cannot lower.",
        "Direct via IB, all-in price.",
    ])

    # ── 5. COLOR (post-award) ──────────────────────────────────────────────
    _cfg_by_id_msg = {c["line_id"]: c for c in _lines_cfg}
    _RATING_ORDER = ["AAA", "AA", "A", "BBB", "BB", "B", "Equity", ""]

    def _rating_idx(r):
        return _RATING_ORDER.index(r) if r in _RATING_ORDER else 99

    _color_lines = [f"*** COLOR — {_bwic_id} ***", ""]
    if _bwic and _phase == "awarded":
        # Group lines by rating
        _by_rating = {}
        for _line in _bwic.lines:
            _cfg = _cfg_by_id_msg.get(_line.line_id, {})
            _rating = _cfg.get("rating", "") or "—"
            _by_rating.setdefault(_rating, []).append((_line, _cfg))

        for _rating in sorted(_by_rating, key=_rating_idx):
            _color_lines.append(f"--- {_rating} ---")
            _covers, _wides, _tights = [], [], []
            for _line, _cfg in _by_rating[_rating]:
                _aw = _line.award
                if _aw is None:
                    continue
                _deal = _cfg.get("deal_name", "")
                _hdr = f"{_deal} {_line.tranche_name} ${_fmt_face(_line.current_face)}"
                if _aw.is_dnt:
                    _color_lines.append(f"  {_hdr}: DNT — {_aw.dnt_reason}")
                    continue
                _best = _line._best_bid_per_dealer()
                _mkt = _cfg.get("market_spread")
                _vs = ""
                if _aw.spread_at_cover is not None and _mkt:
                    _d = _mkt - _aw.spread_at_cover
                    if abs(_d) >= 0.05:
                        _vs = f" ({'+' if _d > 0 else '−'}{abs(_d):.1f} v mkt)"
                _color_lines.append(f"  {_hdr}:")
                if _aw.cover_dealer:
                    _color_lines.append(
                        f"    Cover @ {_aw.cover_price:.4f}"
                        + (f"  (DM {_aw.spread_at_cover:.1f}{_vs})" if _aw.spread_at_cover is not None else "")
                    )
                    _covers.append(_aw.spread_at_cover)
                else:
                    _color_lines.append("    TRADED — no cover (single bid)")
                _color_lines.append(f"    {len(_best)} bids")
            # rating-tier summary
            if _covers:
                _avg = sum(_covers) / len(_covers)
                _color_lines.append(f"  → {_rating} avg cover DM {_avg:.1f} bps")
            _color_lines.append("")
        _color_lines.append("Thanks for participating.")
    else:
        _color_lines.append("(Color available once awarded.)")
    color_msg = "\n".join(_color_lines)

    # ── Phase-aware "active" message label ─────────────────────────────────
    _active_label = {
        "setup":     "Pre-launch announcement",
        "r1_open":   "Round 1 open",
        "r1_closed": "Round 1 closed — advancers",
        "r2_open":   "Round 2 — last & best",
        "awarded":   "Color (post-trade)",
    }.get(_phase, "Announcement")

    _active_msg = {
        "setup":     announce_msg,
        "r1_open":   r1_open_msg,
        "r1_closed": r1_closed_msg,
        "r2_open":   r2_open_msg,
        "awarded":   color_msg,
    }.get(_phase, announce_msg)

    _all_messages_md = mo.accordion({
        "📣 1 · BWIC announcement (pre-launch)": mo.md(f"```text\n{announce_msg}\n```"),
        "🔔 2 · Round 1 open":                    mo.md(f"```text\n{r1_open_msg}\n```"),
        "📋 3 · R1 closed — advancers":           mo.md(f"```text\n{r1_closed_msg}\n```"),
        "🎯 4 · Round 2 last & best":             mo.md(f"```text\n{r2_open_msg}\n```"),
        "📊 5 · Color (post-award)":              mo.md(f"```text\n{color_msg}\n```"),
    })

    mo.vstack([
        mo.md("## 📨 Bloomberg IB messages"),
        mo.callout(
            mo.vstack([
                mo.md(f"**Active message** — _{_active_label}_"),
                mo.md(f"```text\n{_active_msg}\n```"),
            ]),
            kind="success",
        ),
        mo.md("### All messages"),
        _all_messages_md,
    ])
    return


if __name__ == "__main__":
    app.run()
